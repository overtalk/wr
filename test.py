# -*- coding: UTF-8 -*-

import argparse
import openpyxl
import copy
from openpyxl.styles import PatternFill, Alignment, Font, Side, Border, colors
from openpyxl.utils import get_column_letter
from utils import consts

# region ------------- 价格段 -------------
def getPriceSegName(index):
	if index == 0:
		key = '%s+'%(consts.PRICE_SEGMENT[0])
	elif index == len(consts.PRICE_SEGMENT):
		key = '%s-'%(consts.PRICE_SEGMENT[-1])
	else:
		key = '%s-%s'%(consts.PRICE_SEGMENT[index], consts.PRICE_SEGMENT[index-1])
	return key

def getPriceSeg(price):
	# 获取单价属于哪个加个区间
	index = None
	for price_seg_index in xrange(len(consts.PRICE_SEGMENT)):
		if price > consts.PRICE_SEGMENT[price_seg_index]:
			index = price_seg_index
			break

	if index is None:
		return len(consts.PRICE_SEGMENT)

	return index

# endregion ------------- 价格段 -------------

# region ------------- 产品类型 -------------
def getProductType(pt, door_num):

	type_dict = {
		u"T型": [{'door_num':1, 'pt':[u"多门", u'4门']}],
		u'三门': [{'door_num':0, 'pt':[u"三门"]}],
		u'对开': [{'door_num':0, 'pt':[u"对开门"]}],

		u'双门': [
			{'door_num':0, 'pt':[u"双门"]},
			{'door_num':0, 'pt':[u"单门"]}
		],

		u'多门': [
			{'door_num':0, 'pt':[u"多门", u'4门']},
			{'door_num':0, 'pt':[u"多门", u'5门']},
			{'door_num':0, 'pt':[u"多门", u'6门']},
			{'door_num':0, 'pt':[u"多门", u'其它']},
		],
	}

	for type_name, limits in type_dict.iteritems():
		for limit in limits:
			limit_door_num = limit['door_num']
			if door_num != limit_door_num:
				continue

			flag = True
			limit_pt_list= limit['pt']
			for pt_item in limit_pt_list:
				if pt_item not in pt:
					flag = False
					break
			if flag:
				return type_name


	print("%s - %s "%(door_num, pt))
	raise RuntimeError('Unknown Product Type ')
# endregion ------------- 产品类型 -------------

def getExcelData(args):
	# NOTE: 读取excel表格数据
	# 会对重复的数据做融合
	chinese_path = args.excel.decode('utf8').encode('gbk')
	chinese_sheet = args.sheet.decode('utf8').encode('gbk')

	workbook = openpyxl.load_workbook(filename=chinese_path)
	sheet = workbook[chinese_sheet]

	# NOTE: datas 为表中的数据
	datas = tuple(sheet.rows)

	# NOTE: 将表格中的数据读到内存中
	title_name_2_index, title_row_2_name = {}, {}
	raw_datas = {}
	for index in range(len(datas)):
		items = datas[index]

		# 是否是title行
		is_title_flag = True if index == 0 else False
		tmp_raw_data = {}
		for tmp_index in xrange(len(items)):
			item = items[tmp_index]
			value = item.value

			if is_title_flag:
				title_name_2_index[value] = tmp_index
				title_row_2_name[tmp_index] = value
				continue

			title_name = title_row_2_name[tmp_index]
			tmp_raw_data[title_name] = value

		if is_title_flag:
			continue

		# NOTE: 计算份额
		tmp_raw_data[consts.REQUIRED_KEY_SELL_COUNT] = int(tmp_raw_data[consts.REQUIRED_KEY_SELL_COUNT])
		tmp_raw_data[consts.REQUIRED_KEY_SINGLE_PRICE] = int(tmp_raw_data[consts.REQUIRED_KEY_SINGLE_PRICE])
		tmp_raw_data[consts.POST_KEY_SALES] = tmp_raw_data[consts.REQUIRED_KEY_SELL_COUNT]*tmp_raw_data[consts.REQUIRED_KEY_SINGLE_PRICE]
		tmp_raw_data[consts.POST_KEY_NAME] = "%s %s"%(tmp_raw_data[consts.REQUIRED_KEY_CATEGORY], tmp_raw_data[consts.REQUIRED_KEY_MODEL])
		tmp_raw_data[consts.POST_KEY_TIME_DUR] = "%s_%s"%(tmp_raw_data[consts.REQUIRED_KEY_YEAR], tmp_raw_data[consts.REQUIRED_KEY_WEEK])
		tmp_raw_data[consts.POST_KEY_NEW_TYPE] = getProductType(tmp_raw_data[consts.REQUIRED_KEY_PRODECT_TYPE], int(tmp_raw_data[consts.REQUIRED_KEY_IS_MUTI_DOOR]))

		raw_datas[index] = tmp_raw_data

	# NOTE: 检查分类的key是否都存在
	for key in consts.REQUIRED_KEYS:
		is_exist = key in title_name_2_index
		if not is_exist:
			raise RuntimeError("required_key(%s) is absent"%(key))

	# NOTE: 找出重复的行
	repetition_dict = {}
	for index, raw_data in raw_datas.iteritems():
		year = raw_data[consts.REQUIRED_KEY_YEAR]
		week = raw_data[consts.REQUIRED_KEY_WEEK]
		category = raw_data[consts.REQUIRED_KEY_CATEGORY]
		model = raw_data[consts.REQUIRED_KEY_MODEL]
		region = raw_data[consts.REQUIRED_KEY_REGION]
		unique_key = "%s_%s_%s_%s_%s"%(year, week, category, model, region)
		repetition_dict.setdefault(unique_key, []).append(index)

	# NOTE: 处理重复的行
	for unique, index_list in repetition_dict.items():
		if len(index_list) == 1:
			raw_datas[index_list[0]][consts.POST_KEY_MERGE_RABLE_IDS] = index_list
			continue

		# NOTE: 有重复的行
		print("[merge_repeated] unique(%s) index_list(%s)"%(unique, index_list))
		total_sale_count = 0 # 销售量
		total_sales = 0 # 销售额
		for table_id in index_list:
			raw_data = raw_datas[table_id]

			total_sale_count += raw_data[consts.REQUIRED_KEY_SELL_COUNT]
			total_sales += raw_data[consts.POST_KEY_SALES]

		raw_datas[index_list[0]][consts.REQUIRED_KEY_SELL_COUNT] = total_sale_count
		raw_datas[index_list[0]][consts.POST_KEY_SALES] = total_sales
		raw_datas[index_list[0]][consts.REQUIRED_KEY_SINGLE_PRICE] = total_sales/total_sale_count
		raw_datas[index_list[0]][consts.POST_KEY_MERGE_RABLE_IDS] = index_list[1:]

	return raw_datas

def categoryExcelData(year, week, raw_datas):
	# 将excel表中的数据分类

	unique_key = "%s_%s"%(year, week)

	# NOTE: 按照分类的key进行分类
	classified_data = {} # 分类之后的数据
	for talb_id, table_data in raw_datas.iteritems():
		if table_data[consts.POST_KEY_TIME_DUR] != unique_key:
			continue

		for catetory_key in consts.CATEGORY_KEYS:
			price_seg = getPriceSeg(table_data[consts.REQUIRED_KEY_SINGLE_PRICE])
			classified_data \
				.setdefault(table_data[consts.REQUIRED_KEY_REGION], {}) \
				.setdefault(price_seg, {}) \
				.setdefault(catetory_key, {}) \
				.setdefault(table_data[catetory_key], []).append(
					[talb_id, table_data[consts.POST_KEY_SALES]]
				)
	return classified_data

def getMarketSharesForPriceSeg(args, sorted_classified_data, raw_datas):
	# NOTE: 计算每个分段的占比以及
	# 目标品牌在某个价格段所占的份额
	sales_data = {}
	for region, region_data in sorted_classified_data.iteritems():
		# NOTE: 只处理特定地区的
		if region != args.target_region:
			continue

		for price_seg, price_seg_data in region_data.iteritems():
			tmp_total_sales = 0
			tmp_target_sales = 0
			for cat, details in price_seg_data[consts.REQUIRED_KEY_CATEGORY].iteritems():
				for item in details:
					table_id, sales = item
					catgory = raw_datas[table_id][consts.REQUIRED_KEY_CATEGORY]

					tmp_total_sales += sales
					if catgory == args.target_category:
						tmp_target_sales += sales
			sales_data[price_seg] = {
				'total': tmp_total_sales,
				'target': tmp_target_sales,
				'target_market_shares': float(float(tmp_target_sales)/float(tmp_total_sales))
			}

	# NOTE: 计算总的市场销售额
	total_sales = 0
	for price_seg, item in sales_data.iteritems():
		total_sales += item['total']

	ret_dict = {}
	for price_seg, item in sales_data.iteritems():
		ret_dict[price_seg] = {
			'total_market_shares': float(item['total'])/float(total_sales),
			'target_market_shares': item['target_market_shares'],
		}
	return ret_dict

def getMarketSharesForTotal(args, sorted_classified_data, raw_datas,
	required_target_key=None):
	# 整体占比

	processed_required_target_key = required_target_key or consts.CATEGORY_KEYS[0]

	sales_data = []
	for region, region_data in sorted_classified_data.iteritems():

		# NOTE: 只处理特定地区的
		if region != args.target_region:
			continue

		for price_seg, price_seg_data in region_data.iteritems():
			for cat, details in price_seg_data[processed_required_target_key].iteritems():
				for item in details:
					table_id, sales = item
					catgory = raw_datas[table_id][processed_required_target_key]
					sales_data.append([table_id, price_seg, catgory, sales])

	# NOTE: 计算销售总额度
	total_sales = 0
	for item in sales_data:
		total_sales += item[-1]

	# NOTE: 计算份额占比
	for index in xrange(len(sales_data)):
		sales_data[index].append(float(sales_data[index][-1])/float(total_sales))

	# NOTE: 按照份额排序
	sorted_sales_data = sorted(sales_data, key=lambda s: s[-1], reverse=True)

	# NOTE: 取出required_target_key取出需要的数据
	ret_dict = {}
	for index in xrange(len(sorted_sales_data)):
		item = sorted_sales_data[index]
		table_id, price_seg, catgory, sales, percent = item
		raw_data = raw_datas[table_id]
		if required_target_key is not None:
			ret_dict.setdefault(price_seg, {}) \
				.setdefault(catgory, []) \
				.append([table_id, raw_data[consts.POST_KEY_NAME], sales, percent, index+1])
		else:
			ret_dict.setdefault(price_seg, []) \
				.append([table_id, raw_data[consts.POST_KEY_NAME], sales, percent, index+1])

	return ret_dict

class PostData(object):
	def __init__(self, args,
			year, week,
			raw_datas,
			market_shares_for_all_price_seg,
			market_shares_for_total_by_cat,
			market_shares_for_target_by_product_type
		):
		self.args = args
		self.year = year
		self.week = week
		self.raw_datas = raw_datas
		self.market_shares_for_all_price_seg = market_shares_for_all_price_seg
		self.market_shares_for_total_by_cat = market_shares_for_total_by_cat
		self.market_shares_for_target_by_product_type = market_shares_for_target_by_product_type

		self.datas = []

	def post(self):
		self.getAllProductTypes()

		for index in xrange(len(consts.PRICE_SEGMENT)+1):
			self._postOnePriceStage(index)

	def getAllProductTypes(self):
		# 获取所有产品类型
		all_product_type_keys = {}
		for index in xrange(len(consts.PRICE_SEGMENT)+1):
			for product_type, _ in self.market_shares_for_target_by_product_type.get(index, {}).iteritems():
				all_product_type_keys[product_type] = 1
		self.all_product_type_keys = all_product_type_keys.keys()

	def _postOnePriceStage(self, index, top_n=5):
		seg_name = getPriceSegName(index)

		# NOTE: 处理数据头
		title_line1_data = ['价位段', '占比', '份额',
			"整体", "", "", "", "",
			"%s"%(self.args.target_category), "", "", "", "",
		]
		title_line2_data = ['', '', '',"型号", "销量", "价格", "占比" ,"排名","型号", "销量", "价格", "占比" ,"排名"]

		# NOTE: 根据品牌来排序
		market_shares_for_total_by_cat_data = self.market_shares_for_total_by_cat.get(index, {})
		all_data_by_cat = []
		target_catetogy_data_by_cat = []
		for category, product_rank_list in market_shares_for_total_by_cat_data.iteritems():
			all_data_by_cat.extend(product_rank_list)
			if category == self.args.target_category:
				target_catetogy_data_by_cat.extend(product_rank_list)

		# NOTE: 这儿排序的key是排名，所以不用降序
		sorted_all_data_by_cat = sorted(all_data_by_cat, key=lambda rank_index: rank_index[-1])
		sorted_target_catetogy_data_by_cat = sorted(target_catetogy_data_by_cat, key=lambda rank_index: rank_index[-1])

		# NOTE: 根据产品类型来排序
		product_type_data = {}
		market_shares_for_target_by_product_type_data = self.market_shares_for_target_by_product_type.get(index, {})
		for product_type, product_rank_list in market_shares_for_target_by_product_type_data.iteritems():
			product_type_data.setdefault(product_type, []).extend(product_rank_list)

		for product_type in product_type_data.keys():
			data_list = product_type_data[product_type]
			product_type_data[product_type] = sorted(data_list, key=lambda s: s[-1])

		for top_index in xrange(top_n):
			line_data = []

			# NOTE: 处理该价格段位占比
			if top_index == 0:
				data = self.market_shares_for_all_price_seg.get(index, {})
				line_data.extend([
					seg_name,
					data.get('total_market_shares', 0.0),
					data.get('target_market_shares', 0.0),
				])
			else:
				line_data.extend([""]*3)

			# NOTE: 处理整体数据
			if top_index >= len(sorted_all_data_by_cat):
				line_data.extend([""]*5)
			else:
				tmp = sorted_all_data_by_cat[top_index]

				table_id = tmp[0]
				product_name = tmp[1]
				percent = tmp[-2]
				rank = tmp[-1]

				table_data = raw_datas[table_id]
				line_data.extend([
					product_name,
					table_data[consts.REQUIRED_KEY_SELL_COUNT],
					table_data[consts.REQUIRED_KEY_SINGLE_PRICE],
					percent,
					rank,
				])

			# NOTE: 处理目标品牌的数据
			if top_index >= len(sorted_target_catetogy_data_by_cat):
				line_data.extend([""]*5)
			else:
				tmp = sorted_target_catetogy_data_by_cat[top_index]

				table_id = tmp[0]
				product_name = tmp[1]
				percent = tmp[-2]
				rank = tmp[-1]

				table_data = raw_datas[table_id]
				line_data.extend([
					product_name,
					table_data[consts.REQUIRED_KEY_SELL_COUNT],
					table_data[consts.REQUIRED_KEY_SINGLE_PRICE],
					percent,
					rank,
				])

			# NOTE: 处理各种品牌
			for product_type in self.all_product_type_keys:
				data_list = product_type_data.get(product_type, [])

				# NOTE: 增加title
				if top_index == 0:
					title_line1_data.extend(["%s"%(product_type), "", "", "", "",])
					title_line2_data.extend(["型号", "销量", "价格", "占比" ,"排名"])

				if top_index >= len(data_list):
					line_data.extend([""]*5)
				else:
					tmp = data_list[top_index]

					table_id = tmp[0]
					product_name = tmp[1]
					percent = tmp[-2]
					rank = tmp[-1]

					table_data = raw_datas[table_id]
					line_data.extend([
						product_name,
						table_data[consts.REQUIRED_KEY_SELL_COUNT],
						table_data[consts.REQUIRED_KEY_SINGLE_PRICE],
						percent,
						rank,
					])

			# NOTE: 将行数据存储下来
			if top_index == 0 and index == 0:
				self.datas.append(title_line1_data)
				self.datas.append(title_line2_data)
			self.datas.append(line_data)

	def saveToExcel(self):
		workbook = openpyxl.Workbook()
		new_sheet = workbook.active
		new_sheet.title = 'test_sheet'

		for data in self.datas:
			new_sheet.append(data)
		self.addTableFormat(new_sheet)

		workbook.save('%s_%s.xlsx'%(self.year, self.week))

	# region ---------------- excel 表格式相关 ----------------
	def addTableFormat(self, new_sheet):

		max_row = new_sheet.max_row
		max_column = new_sheet.max_column

		if max_column % 5 != 3:
			raise RuntimeError("xxxxxxxxxx")

		# NOTE: 合并单元格 & 填充颜色
		color = PatternFill("solid", fgColor="1874CD")
		alignment = Alignment(horizontal='center', vertical='center')
		font = Font(name=u'微软雅黑', size=10)

		# NOTE: 所有的单元格设置成居中
		for table_row in xrange(1, max_row+1):
			for table_column in xrange(1, max_column+1):
				cell = new_sheet.cell(row=table_row, column=table_column)
				cell.font = font
				cell.alignment = alignment

		# NOTE: 设置表格头的底色
		for i in xrange(3):
			new_sheet.merge_cells(start_row=1, start_column=i+1, end_row=2, end_column=i+1)
			cell = new_sheet.cell(row=1, column=i+1)
			cell.fill = color # 设置单元格颜色

		# NOTE: 合并单元格
		column_step = [1]
		for i in xrange(max_column/5):
			start = i*5+4
			end = (i+1)*5+3
			column_step.append(start)

			new_sheet.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
			cell = new_sheet.cell(row=1, column=start)
			cell.fill = color

			for column in xrange(start, end+1):
				new_sheet.cell(row=2, column=column).fill = color
		column_step.append(max_column+1)

		# NOTE: 自适应列宽
		# 第一步：计算每列最大宽度，并存储在列表lks中。
		lks = [] #英文变量太费劲，用汉语首字拼音代替
		for i in range(1, max_column+1): #每列循环
			lk = 1 #定义初始列宽，并在每个行循环完成后重置
			for j in range(1, max_row + 1): #每行循环
				sz = new_sheet.cell(row=j,column=i).value #每个单元格内容
				if isinstance(sz,str): #中文占用多个字节，需要分开处理
					lk1 = len(sz.encode('gbk')) #gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
				else:
					lk1 = len("%s"%(sz))
				if lk < lk1:
					lk = lk1 #借助每行循环将最大值存入lk中
				# print(lk)
			lks.append(lk) # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）

		# 第二步：设置列宽
		for i in range(1, max_column +1):
			k = get_column_letter(i) #将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
			new_sheet.column_dimensions[k].width = lks[i-1]+2 #设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整

		# NOTE: 边框
		for row in xrange(len(consts.PRICE_SEGMENT)+1):
			s_row = 3+row*5
			e_row = s_row+4
			for tmp_column_index in xrange(1, len(column_step)):
				s_column = column_step[tmp_column_index-1]
				e_column = column_step[tmp_column_index]-1
				self.addBorder(new_sheet, s_row, e_row, s_column, e_column)


	def addBorder(self, new_sheet, s_row, e_row, s_column, e_column):
		#设置左粗框线
		for row_index in xrange(s_row, e_row+1):
			cell = new_sheet.cell(row=row_index, column=s_column)
			cell.border = self.my_border(cell.border.top.style, cell.border.bottom.style,
				'medium', cell.border.right.style)

		# 设置右粗框线
		for row_index in xrange(s_row, e_row+1):
			cell = new_sheet.cell(row=row_index, column=e_column)
			cell.border = self.my_border(cell.border.top.style, cell.border.bottom.style,
				cell.border.left.style, 'medium')

		# 设置上粗框线
		for column_index in xrange(s_column, e_column+1):
			cell = new_sheet.cell(row=s_row, column=column_index)
			cell.border = self.my_border('medium', cell.border.bottom.style,
				cell.border.left.style, cell.border.right.style)

		# 设置下粗框线
		for column_index in xrange(s_column, e_column+1):
			cell = new_sheet.cell(row=e_row, column=column_index)
			cell.border = self.my_border(cell.border.top.style, 'medium',
					cell.border.left.style, cell.border.right.style)

	def my_border(self, t_border, b_border, l_border, r_border):
		border = Border(top=Side(border_style=t_border, color=colors.BLACK),
						bottom=Side(border_style=b_border, color=colors.BLACK),
						left=Side(border_style=l_border, color=colors.BLACK),
						right=Side(border_style=r_border, color=colors.BLACK))
		return border
	# endregion ---------------- excel 表格式相关 ----------------

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument("--excel", type=str, required=False,
				default='D:\python\wr\新建文件夹\wr1.xlsx', help='excel path')
	parser.add_argument("--sheet", type=str, required=False,
				default='Sheet1', help='sheet name')
	parser.add_argument("--target_region", type=unicode, required=False,
				default=u'杭州', help='target city')
	parser.add_argument("--target_category", type=unicode, required=False,
				default=u'海尔', help='target category')

	# 年份 & 周
	parser.add_argument("--target_year", type=int, required=True,  help='target year')
	parser.add_argument("--target_week", type=int, required=True,  help='target week')

	args = parser.parse_args()

	chinese_path = args.excel.decode('utf8').encode('gbk')
	chinese_sheet = args.sheet.decode('utf8').encode('gbk')
	print("[ARGS] excel(%s) - sheet(%s)" % (chinese_path, chinese_sheet))


	t = args.excel.decode('utf8').encode('gbk')

	# NOTE: 读取excel中的数据
	raw_datas = getExcelData(args)

	# NOTE: 按照分类的key进行分类
	classified_data = categoryExcelData(args.target_year, args.target_week, raw_datas) # 分类之后的数据

	# NOTE: 检查是否有重复的数据
	duplicates_dict = {}
	for region, region_data in classified_data.iteritems():
		for price_seg, price_seg_data in region_data.iteritems():
			for cat_key in consts.CATEGORY_KEYS:
				tmp_duplicates_dict = {}
				for cat, details in price_seg_data[cat_key].iteritems():
					for item in details:
						table_id = item[0]
						raw_data = raw_datas[table_id]

						# NOTE: 获取去重的key
						region_value = raw_data[consts.REQUIRED_KEY_REGION]
						catgoty_value = raw_data[consts.REQUIRED_KEY_CATEGORY]
						model_value = raw_data[consts.REQUIRED_KEY_MODEL]
						model_value = raw_data[consts.POST_KEY_TIME_DUR]
						unique_key = "%s_%s_%s_%s"%(model_value, region_value, catgoty_value, model_value)

						tmp_duplicates_dict.setdefault(unique_key, []).append(table_id)
				duplicates_dict[cat_key] = copy.deepcopy(tmp_duplicates_dict)

	for k1, v1 in duplicates_dict.iteritems():
		for k2, v2 in v1.iteritems():
			if len(v2) > 1:
				print("[DUPLICATED] cat_key(%s) unique_id(%s) row_ids(%s)"%
					(k1, k2, v2))

	# NOTE: 排序
	sorted_classified_data = {} # 排序之后的数据
	for region, region_data in classified_data.iteritems():
		for price_seg, price_seg_data in region_data.iteritems():
			for cat_key in consts.CATEGORY_KEYS:
				for cat, details in price_seg_data[cat_key].iteritems():
					sorted_data = sorted(details, key=lambda s: s[1], reverse=True)
					sorted_classified_data.setdefault(region, {}) \
						.setdefault(price_seg, {}) \
						.setdefault(cat_key, {})[cat] = sorted_data

	# NOTE: 每个加个段占市场的总份额 & 目标品牌在每个段的占比
	market_shares_for_all_price_seg = getMarketSharesForPriceSeg(args, sorted_classified_data, raw_datas)

	# NOTE: 整体对比
	market_shares_for_total_by_cat = getMarketSharesForTotal(args, sorted_classified_data, raw_datas,
		required_target_key=consts.REQUIRED_KEY_CATEGORY)

	# NOTE: 目标品牌
	market_shares_for_target_by_product_type = getMarketSharesForTotal(args, sorted_classified_data, raw_datas,
		required_target_key=consts.POST_KEY_NEW_TYPE)

	# NOTE: 将结果写入excel中
	post_data_util = PostData(
		args, args.target_year, args.target_week,
		raw_datas,
		market_shares_for_all_price_seg,
		market_shares_for_total_by_cat,
		market_shares_for_target_by_product_type)
	post_data_util.post()
	post_data_util.saveToExcel()

