# -*- coding: UTF-8 -*-

import openpyxl

def test_load_excel():
    workbook = openpyxl.load_workbook(filename="test.xlsx")
    # 2.通过 sheet 名称获取表格
    sheet = workbook["Sheet1"]
    print(sheet)


    # 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
    print(sheet.dimensions)
    # 4.获取表格内某个格子的数据
    # 1 sheet["A1"]方式
    cell1 = sheet["A1"]
    cell2 = sheet["C11"]
    print(cell1.value, cell2.value)
    """
	workbook.active 打开激活的表格; sheet["A1"] 获取 A1 格子的数据; cell.value 获取格子中的值;
	"""
    # 4.2sheet.cell(row=, column=)方式
    cell1 = sheet.cell(row = 1,column = 1)
    cell2 = sheet.cell(row = 11,column = 3)
    print(cell1.value, cell2.value)

    # 5. 获取一系列格子
    # 获取 A1:C2 区域的值
    cell = sheet["A1:C2"]
    print(cell)
    for i in cell:
        for j in i:
            print(j.value)

def test_write_excel():
    workbook = openpyxl.Workbook()

    new_sheet = workbook.active
    new_sheet.title = 'test_sheet'
    new_sheet['A1'] = '国家'
    new_sheet['B1'] = '首都'

    data = {
        '国家1':"首都1",
        '国家2':'首都1',
        '国家3':'首都1',
        '国家4':'首都1',
    }

    for k in data:
        new_sheet.append([k,data[k]])

    workbook.save('test1.xlsx')

if __name__ == "__main__":
    test_load_excel()
    test_write_excel()
